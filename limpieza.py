import pandas as pd
from openpyxl import load_workbook

# Cargar el archivo de Excel
file_path = r"insertar ruta del archivo"
workbook = load_workbook(file_path)
sheet_names = workbook.sheetnames

# Crear un nuevo archivo de Excel para guardar los resultados
output_file = "nombre_archivo_procesado.xlsx"
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Procesar cada hoja en el archivo
for sheet_name in sheet_names:
    # Leer la hoja actual
    data = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=3)

    # Eliminar las primeras 4 columnas
    data = data.iloc[:, 4:]

    # Función para dividir las celdas basándose en saltos de página
    def split_cell(cell):
        if "\n" in str(cell):
            return cell.split("\n")
        else:
            return [cell, None, None]

    # Aplicar la función de división solo a las columnas L en adelante
    for col in data.columns[7:]:
        new_cols = data[col].apply(split_cell).apply(pd.Series)
        new_cols.columns = [f"{col}_1", f"{col}_2", f"{col}_3"]
        data = pd.concat([data, new_cols], axis=1)
        data.drop(col, axis=1, inplace=True)

    # Guardar la hoja procesada en el nuevo archivo de Excel
    data.to_excel(writer, sheet_name=sheet_name, index=False)

# Guardar y cerrar el archivo de Excel
writer.save()
print(f"Datos procesados y guardados en {output_file}")
